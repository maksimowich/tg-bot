from datetime import datetime
import os
from typing import Sequence
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from src.core.config import app_settings
from src.db import CartItemOrm


ORDERS_TITLE = [
    "Order ID",
    "Creation Datetime",
    "User ID",
    "Address",
    "Payment Amount",
    "Payment Link",
]

ORDER_ITEMS_TITLE = [
    "Order ID",
    "Product ID",
    "Quantity",
]


async def save_order(
        order_id: UUID,
        creation_dttm: datetime,
        payment_link: str,
        user_id: str,
        address: str,
        payment_amount: float,
        cart_items: Sequence[CartItemOrm],
) -> UUID:
    if os.path.exists(app_settings.excel_filepath):
        workbook = load_workbook(app_settings.excel_filepath)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    if "Orders" in workbook.sheetnames:
        orders_sheet = workbook["Orders"]
    else:
        orders_sheet = workbook.create_sheet("Orders")
        orders_sheet.append(ORDERS_TITLE)
        for cell in orders_sheet[1]:
            cell.font = Font(bold=True)

    if "Order Items" in workbook.sheetnames:
        items_sheet = workbook["Order Items"]
    else:
        items_sheet = workbook.create_sheet("Order Items")
        items_sheet.append(ORDER_ITEMS_TITLE)
        for cell in items_sheet[1]:
            cell.font = Font(bold=True)

    creation_dttm_str = creation_dttm.strftime("%Y-%m-%d %H:%M:%S")

    orders_sheet.append([
        str(order_id),
        creation_dttm_str,
        user_id,
        address,
        payment_amount,
        payment_link,
    ])

    for cart_item in cart_items:
        items_sheet.append([
            str(order_id),
            str(cart_item.product_id),
            cart_item.quantity,
        ])

    workbook.save(app_settings.excel_filepath)
    return order_id
